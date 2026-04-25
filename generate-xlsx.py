#!/usr/bin/env python3
"""
Embafinans BA Practice Artifacts - XLSX Generator
Generates 12 professionally formatted XLSX files using openpyxl.
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle
)

# ── Constants ────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2A6496", end_color="2A6496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=10, color="333333")
ALT_FILL = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)

BASE_DIR = "/home/z/my-project/embafinans"


# ── Helper Functions ─────────────────────────────────────────────────────────

def create_workbook(sheet_name):
    """Create a new workbook with a single sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws


def style_header_cell(cell):
    """Apply header styling to a cell."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_data_cell(cell, row_idx, center=False):
    """Apply data styling to a cell with alternating row colors."""
    cell.font = DATA_FONT
    cell.fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN if center else WRAP_ALIGN


def add_header_row(ws, headers, col_widths=None):
    """Add a styled header row and set column widths."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell)
    # Set column widths
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(w, 15)
    else:
        for col_idx in range(1, len(headers) + 1):
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = 22
    # Freeze pane
    ws.freeze_panes = "A2"


def add_data_row(ws, row_data, row_num, center_cols=None):
    """Add a styled data row. center_cols is a set of 1-based column indices to center."""
    if center_cols is None:
        center_cols = set()
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        style_data_cell(cell, row_num, center=(col_idx in center_cols))


def save_workbook(wb, filepath):
    """Save workbook and print status."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"  ✓ Created: {filepath}")


# ── File 1: User Stories – Credit Scoring ───────────────────────────────────

def gen_file_01():
    path = os.path.join(BASE_DIR, "01-credit-scoring", "User_Stories_Credit_Scoring.xlsx")
    wb, ws = create_workbook("User Stories")
    headers = [
        "ID", "Epic", "As a (Role)", "I want to (Action)",
        "So that (Benefit)", "Acceptance Criteria (Given/When/Then)",
        "Priority", "Story Points", "Status"
    ]
    col_widths = [10, 20, 22, 30, 30, 50, 12, 14, 12]
    add_header_row(ws, headers, col_widths)

    stories = [
        [
            "US-101", "Credit Application",
            "Customer",
            "submit a credit application via mobile app",
            "I can apply for credit without visiting a branch",
            "Given I am a registered customer on the app\nWhen I fill in my personal and financial details and submit\nThen my application should be saved and sent for scoring",
            "High", 8, "Ready"
        ],
        [
            "US-102", "Scoring Engine",
            "Risk Analyst",
            "view automated scoring results",
            "I can focus on manual review cases only",
            "Given an application has been scored\nWhen the score is between 50-79\nThen the application appears in my manual review queue",
            "High", 5, "Ready"
        ],
        [
            "US-103", "Notifications",
            "Customer",
            "receive SMS notification with credit decision",
            "I know the result immediately",
            "Given the scoring engine returns a decision\nWhen the decision is AUTO_APPROVED\nThen an SMS is sent to the customer within 30 seconds",
            "High", 3, "Ready"
        ],
        [
            "US-104", "Scoring Model",
            "Risk Manager",
            "configure scoring factor weights",
            "I can adjust the scoring model",
            "Given I am logged in as risk manager\nWhen I change the weight of any scoring factor and save\nThen the new weights apply to all future scoring requests",
            "Medium", 5, "Ready"
        ],
        [
            "US-105", "Manual Review",
            "Credit Committee Member",
            "see applications flagged for manual review with full details",
            "I can make informed approval decisions",
            "Given an application has score 50-79\nWhen I open the manual review queue\nThen I see applicant data, bureau score, and scoring factor breakdown",
            "High", 8, "Ready"
        ],
        [
            "US-106", "Monitoring",
            "System Admin",
            "view scoring performance dashboard",
            "I can monitor system health",
            "Given the scoring dashboard is accessed\nWhen I view the main page\nThen I see total applications, approval rate, average processing time, and error count for today",
            "Medium", 5, "Ready"
        ],
        [
            "US-107", "Credit Application",
            "Customer",
            "the system to pre-fill my data from previous applications",
            "I don't have to re-enter information",
            "Given I have submitted a previous application\nWhen I start a new application\nThen my personal data is pre-filled and editable",
            "Low", 3, "Backlog"
        ],
        [
            "US-108", "Partner Integration",
            "Partner Store",
            "check if a customer is pre-approved for credit",
            "I can offer credit at point of sale",
            "Given a partner store agent enters customer PIN\nWhen the pre-screen API is called\nThen the response shows eligible/ineligible with pre-approved limit",
            "High", 8, "Ready"
        ],
        [
            "US-109", "Reporting",
            "Risk Analyst",
            "export scoring results to Excel",
            "I can perform custom analysis",
            "Given I am on the scoring dashboard\nWhen I click export and select date range\nThen a CSV file downloads with all scoring results",
            "Low", 2, "Backlog"
        ],
        [
            "US-110", "Customer Experience",
            "Customer",
            "see my credit scoring factors explanation",
            "I understand why I was approved or rejected",
            "Given the scoring decision is AUTO_REJECTED\nWhen I view the decision details\nThen I see a breakdown of all scoring factors with my values",
            "Medium", 5, "Ready"
        ],
        [
            "US-111", "Scoring Engine",
            "Risk Analyst",
            "the system to handle bureau API timeouts gracefully",
            "applications are not lost when external services are slow",
            "Given the credit bureau API is unavailable\nWhen a scoring request is submitted\nThen the application is queued for retry and the applicant is notified of processing delay",
            "High", 5, "Ready"
        ],
        [
            "US-112", "Scoring Model",
            "Risk Manager",
            "view scoring model version history",
            "I can track changes and revert if needed",
            "Given scoring factor weights have been changed\nWhen I view the version history\nThen I see all previous configurations with timestamps and changed-by information",
            "Low", 3, "Backlog"
        ],
    ]
    center = {1, 7, 8, 9}
    for idx, story in enumerate(stories):
        add_data_row(ws, story, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 2: RICE – Credit Scoring ───────────────────────────────────────────

def gen_file_02():
    path = os.path.join(BASE_DIR, "01-credit-scoring", "RICE_Credit_Scoring.xlsx")
    wb, ws = create_workbook("RICE Prioritization")
    headers = [
        "ID", "Requirement", "Reach (users/month)", "Impact (1-3)",
        "Confidence (%)", "Effort (person-weeks)", "RICE Score (R×I×C/E)", "Priority Rank"
    ]
    col_widths = [10, 35, 20, 14, 16, 22, 24, 14]
    add_header_row(ws, headers, col_widths)

    reqs = [
        ["REQ-101", "Automated scoring engine", 400, 3, 0.90, 8],
        ["REQ-102", "Credit bureau API integration", 400, 3, 0.85, 3],
        ["REQ-103", "SMS notification service", 400, 2, 0.95, 1],
        ["REQ-104", "Scoring dashboard", 15, 2, 0.80, 4],
        ["REQ-105", "Pre-screen API for partners", 50, 2, 0.75, 2],
        ["REQ-106", "Manual review queue", 125, 3, 0.85, 3],
        ["REQ-107", "Scoring factor weight configuration", 5, 2, 0.90, 2],
        ["REQ-108", "Export to CSV functionality", 15, 1, 0.95, 0.5],
        ["REQ-109", "Customer data pre-fill", 400, 1, 0.80, 1],
        ["REQ-110", "Scoring explanation for customers", 400, 2, 0.70, 3],
    ]
    # Calculate RICE scores
    for r in reqs:
        rice = (r[2] * r[3] * r[4]) / r[5]
        r.append(round(rice, 1))
    # Sort by RICE descending to assign ranks
    ranked = sorted(reqs, key=lambda x: x[6], reverse=True)
    for i, r in enumerate(ranked, 1):
        r.append(i)

    center = {1, 3, 4, 5, 6, 7, 8}
    for idx, row in enumerate(ranked):
        add_data_row(ws, row, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 3: UAT Test Plan – Credit Scoring ──────────────────────────────────

def gen_file_03():
    path = os.path.join(BASE_DIR, "01-credit-scoring", "UAT_Test_Plan_Credit_Scoring.xlsx")
    wb, ws = create_workbook("UAT Test Cases")
    headers = [
        "TC-ID", "Requirement Ref", "Test Scenario", "Pre-conditions",
        "Test Steps", "Expected Result", "Priority", "Status"
    ]
    col_widths = [10, 18, 28, 28, 40, 40, 12, 12]
    add_header_row(ws, headers, col_widths)

    cases = [
        ["TC-101", "REQ-101", "Happy path auto-approval (score >= 80)",
         "Customer is registered; valid personal and financial data available",
         "1. Login to mobile app\n2. Navigate to Credit Application\n3. Fill all required fields\n4. Submit application",
         "Application scored >= 80; status set to AUTO_APPROVED; SMS sent; customer notified within 30 seconds",
         "Critical", "Not Run"],
        ["TC-102", "REQ-101", "Manual review routing (score 50-79)",
         "Customer with borderline credit profile",
         "1. Submit application with score-yielding data 50-79\n2. Check risk analyst queue",
         "Application appears in manual review queue with score and all factor details visible",
         "Critical", "Not Run"],
        ["TC-103", "REQ-101", "Auto-rejection (score < 50)",
         "Customer with poor credit profile",
         "1. Submit application yielding score < 50\n2. Check application status",
         "Application status set to AUTO_REJECTED; rejection SMS sent with explanation link",
         "Critical", "Not Run"],
        ["TC-104", "REQ-105", "Pre-screen eligible customer",
         "Customer has bureau score >= 700; partner store agent authenticated",
         "1. Agent enters customer PIN\n2. System calls pre-screen API\n3. Response received",
         "API returns eligible=true with pre-approved limit amount",
         "High", "Not Run"],
        ["TC-105", "REQ-105", "Pre-screen ineligible customer",
         "Customer has bureau score < 600; partner store agent authenticated",
         "1. Agent enters customer PIN\n2. System calls pre-screen API\n3. Response received",
         "API returns eligible=false with reason code",
         "High", "Not Run"],
        ["TC-106", "REQ-103", "SMS notification sent on approval",
         "Application approved; SMS gateway configured",
         "1. Application auto-approved\n2. Monitor SMS delivery logs",
         "SMS delivered to customer mobile within 30 seconds of approval",
         "High", "Not Run"],
        ["TC-107", "REQ-103", "SMS notification sent on rejection",
         "Application rejected; SMS gateway configured",
         "1. Application auto-rejected\n2. Monitor SMS delivery logs",
         "SMS delivered to customer mobile within 30 seconds of rejection",
         "High", "Not Run"],
        ["TC-108", "REQ-104", "Scoring dashboard shows correct metrics",
         "At least 10 applications processed today; admin logged in",
         "1. Navigate to scoring dashboard\n2. Verify metrics displayed",
         "Total applications, approval rate, average processing time, and error count displayed correctly for today",
         "Medium", "Not Run"],
        ["TC-109", "REQ-102", "Bureau API timeout handling",
         "Bureau API configured to simulate timeout after 5 seconds",
         "1. Submit application\n2. Bureau API times out\n3. Check application status",
         "Application queued for retry; customer notified of processing delay; no data loss",
         "Critical", "Not Run"],
        ["TC-110", "REQ-101", "Invalid application data validation",
         "Customer logged in; mandatory fields available",
         "1. Submit application with missing required fields\n2. Submit with invalid phone number\n3. Submit with negative income",
         "Appropriate validation error messages displayed for each invalid field; application not submitted",
         "High", "Not Run"],
        ["TC-111", "REQ-101", "Duplicate application detection",
         "Customer has pending application in last 30 days",
         "1. Customer attempts to submit a new application\n2. System checks for duplicates",
         "System warns about existing pending application; offers to view status instead of creating duplicate",
         "Medium", "Not Run"],
        ["TC-112", "REQ-107", "Scoring factor weight change",
         "Risk manager logged in; current weights displayed",
         "1. Navigate to scoring factor configuration\n2. Change weight of 'Debt-to-Income Ratio' from 20% to 25%\n3. Save\n4. Submit test application",
         "New weights saved with version record; test application scored using updated weights",
         "High", "Not Run"],
        ["TC-113", "REQ-108", "Export scoring results",
         "At least 20 scoring records exist for selected date range",
         "1. Navigate to scoring dashboard\n2. Click Export\n3. Select date range\n4. Download file",
         "CSV file downloaded with all scoring results for the date range; columns match dashboard fields",
         "Medium", "Not Run"],
        ["TC-114", "REQ-106", "Manual review queue display",
         "At least 5 applications in manual review queue",
         "1. Login as credit committee member\n2. Open manual review queue\n3. Click on first application",
         "All 5 applications listed with score, applicant name, and timestamp; detail view shows full factor breakdown",
         "High", "Not Run"],
        ["TC-115", "REQ-110", "Customer scoring explanation",
         "Customer application auto-rejected; explanation module enabled",
         "1. Customer views rejection notification\n2. Clicks 'View Details'\n3. Reviews scoring factor breakdown",
         "All scoring factors displayed with customer's actual values, weight percentages, and improvement suggestions",
         "Medium", "Not Run"],
    ]
    center = {1, 7, 8}
    for idx, case in enumerate(cases):
        add_data_row(ws, case, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 4: User Stories – B2C Sales Channel ────────────────────────────────

def gen_file_04():
    path = os.path.join(BASE_DIR, "02-b2c-sales-channel", "User_Stories_B2C.xlsx")
    wb, ws = create_workbook("User Stories")
    headers = [
        "ID", "Epic", "As a (Role)", "I want to (Action)",
        "So that (Benefit)", "Acceptance Criteria (Given/When/Then)",
        "Priority", "Story Points", "Status"
    ]
    col_widths = [10, 20, 22, 30, 30, 50, 12, 14, 12]
    add_header_row(ws, headers, col_widths)

    stories = [
        ["US-201", "Product Catalog", "Customer",
         "browse products with filters and categories",
         "I can easily find products I want to purchase",
         "Given I am on the product catalog page\nWhen I apply filters for category, price range, and brand\nThen only matching products are displayed with images and prices",
         "High", 5, "Ready"],
        ["US-202", "Order Management", "Customer",
         "create a new order by adding products to cart",
         "I can select multiple items for purchase",
         "Given I have browsed products\nWhen I add items to cart and proceed to checkout\nThen I see order summary with product details, quantities, and total amount",
         "High", 8, "Ready"],
        ["US-203", "Payment", "Customer",
         "select installment plan options before payment",
         "I can choose a repayment plan that fits my budget",
         "Given I am at checkout with order total calculated\nWhen I view installment options\nThen I see 3, 6, 9, 12, and 18 month plans with monthly amounts and total interest",
         "High", 5, "Ready"],
        ["US-204", "Credit Application", "Customer",
         "apply for credit directly within the order flow",
         "I can finance my purchase seamlessly",
         "Given I selected an installment plan\nWhen I submit credit application with required documents\nThen the system integrates with scoring engine and returns decision in real-time",
         "High", 13, "Ready"],
        ["US-205", "Payment", "Customer",
         "complete payment via mobile money or bank card",
         "I can pay my down payment and processing fees",
         "Given credit is approved\nWhen I select payment method and enter details\nThen payment is processed and order status changes to 'Payment Confirmed'",
         "High", 8, "Ready"],
        ["US-206", "Order Tracking", "Customer",
         "track my order status from approval to delivery",
         "I know when to expect my products",
         "Given I have a confirmed order\nWhen I open order details\nThen I see current status, estimated delivery date, and timeline of status changes",
         "Medium", 5, "Ready"],
        ["US-207", "Payment", "Customer",
         "receive payment confirmation with receipt",
         "I have proof of my payment",
         "Given payment is successfully processed\nWhen the transaction completes\nThen I receive SMS and in-app notification with payment receipt including amount and reference number",
         "Medium", 3, "Ready"],
        ["US-208", "Repayment", "Customer",
         "view my installment schedule and upcoming payments",
         "I can plan my finances accordingly",
         "Given I have an active installment plan\nWhen I navigate to 'My Installments'\nThen I see full schedule with due dates, amounts, paid/upcoming status, and total remaining balance",
         "Medium", 5, "Ready"],
        ["US-209", "Order Management", "Customer",
         "view my complete order history",
         "I can reference past purchases",
         "Given I have placed multiple orders\nWhen I open 'Order History'\nThen I see all orders with dates, statuses, amounts, and ability to filter by status and date range",
         "Low", 3, "Backlog"],
        ["US-210", "Returns", "Customer",
         "request a refund or return for my order",
         "I am protected if the product has issues",
         "Given I have a delivered order within return window\nWhen I select 'Request Return' and provide reason\nThen return request is created and customer service is notified",
         "Low", 5, "Backlog"],
    ]
    center = {1, 7, 8, 9}
    for idx, story in enumerate(stories):
        add_data_row(ws, story, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 5: RICE – B2C Sales Channel ────────────────────────────────────────

def gen_file_05():
    path = os.path.join(BASE_DIR, "02-b2c-sales-channel", "RICE_B2C.xlsx")
    wb, ws = create_workbook("RICE Prioritization")
    headers = [
        "ID", "Requirement", "Reach (users/month)", "Impact (1-3)",
        "Confidence (%)", "Effort (person-weeks)", "RICE Score (R×I×C/E)", "Priority Rank"
    ]
    col_widths = [10, 35, 20, 14, 16, 22, 24, 14]
    add_header_row(ws, headers, col_widths)

    reqs = [
        ["REQ-201", "Mobile app order flow", 1000, 3, 0.90, 10],
        ["REQ-202", "Payment gateway integration", 800, 3, 0.85, 6],
        ["REQ-203", "Installment calculator", 600, 2, 0.90, 2],
        ["REQ-204", "Order tracking system", 500, 2, 0.85, 4],
        ["REQ-205", "SMS notifications", 800, 1, 0.95, 1],
        ["REQ-206", "Installment schedule management", 400, 2, 0.80, 3],
        ["REQ-207", "Order management portal", 200, 1, 0.85, 3],
    ]
    for r in reqs:
        rice = (r[2] * r[3] * r[4]) / r[5]
        r.append(round(rice, 1))
    ranked = sorted(reqs, key=lambda x: x[6], reverse=True)
    for i, r in enumerate(ranked, 1):
        r.append(i)

    center = {1, 3, 4, 5, 6, 7, 8}
    for idx, row in enumerate(ranked):
        add_data_row(ws, row, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 6: UAT Test Plan – B2C Sales Channel ───────────────────────────────

def gen_file_06():
    path = os.path.join(BASE_DIR, "02-b2c-sales-channel", "UAT_Test_Plan_B2C.xlsx")
    wb, ws = create_workbook("UAT Test Cases")
    headers = [
        "TC-ID", "Requirement Ref", "Test Scenario", "Pre-conditions",
        "Test Steps", "Expected Result", "Priority", "Status"
    ]
    col_widths = [10, 18, 28, 28, 40, 40, 12, 12]
    add_header_row(ws, headers, col_widths)

    cases = [
        ["TC-201", "REQ-201", "Create new order with products",
         "Customer logged in; products available in catalog",
         "1. Browse products\n2. Add 3 items to cart\n3. Proceed to checkout\n4. Verify order summary",
         "Order created with correct products, quantities, and calculated total amount",
         "Critical", "Not Run"],
        ["TC-202", "REQ-202", "Initiate payment via mobile money",
         "Order created; customer has mobile money account",
         "1. Select mobile money payment\n2. Enter phone number\n3. Confirm payment\n4. Enter OTP",
         "Payment initiated; OTP sent; on confirmation order status updates to Payment Confirmed",
         "Critical", "Not Run"],
        ["TC-203", "REQ-202", "Handle payment success callback",
         "Payment initiated; mock success callback from payment gateway",
         "1. Initiate payment\n2. Gateway sends success callback\n3. System processes callback",
         "Order status updated to 'Confirmed'; payment record created; confirmation SMS sent",
         "Critical", "Not Run"],
        ["TC-204", "REQ-202", "Handle payment failure callback",
         "Payment initiated; mock failure callback from payment gateway",
         "1. Initiate payment\n2. Gateway sends failure callback\n3. System processes callback",
         "Order status updated to 'Payment Failed'; customer notified to retry; retry option available",
         "Critical", "Not Run"],
        ["TC-205", "REQ-204", "Order status update throughout lifecycle",
         "Order in 'Confirmed' state",
         "1. Simulate status transitions: Confirmed → Processing → Shipped → Delivered\n2. Check customer view after each",
         "Customer sees updated status and timeline after each transition; SMS sent for key status changes",
         "High", "Not Run"],
        ["TC-206", "REQ-203", "Installment schedule generation after approval",
         "Credit approved with 12-month installment plan",
         "1. Credit application approved\n2. Check installment schedule\n3. Verify amounts and dates",
         "12 installment records created with correct monthly amounts, due dates, and total matching order amount plus interest",
         "High", "Not Run"],
        ["TC-207", "REQ-205", "SMS notification on order confirmation",
         "Order payment confirmed; SMS gateway configured",
         "1. Complete order payment\n2. Monitor SMS delivery",
         "SMS sent within 30 seconds with order number, total amount, and estimated delivery",
         "High", "Not Run"],
        ["TC-208", "REQ-206", "View installment schedule and payment history",
         "Customer has 3 active installment plans; 2 payments made",
         "1. Navigate to My Installments\n2. View schedule for each plan\n3. Check payment history",
         "All plans displayed; paid installments marked; upcoming due dates highlighted; remaining balance shown",
         "Medium", "Not Run"],
        ["TC-209", "REQ-201", "Concurrent order handling",
         "Customer logged in on two devices",
         "1. Create order on Device A\n2. Simultaneously add different items on Device B\n3. Submit from both devices",
         "First submission succeeds; second receives conflict error with option to refresh cart",
         "High", "Not Run"],
        ["TC-210", "REQ-201", "Empty cart checkout prevention",
         "Customer has empty cart",
         "1. Navigate directly to checkout URL\n2. Attempt to proceed",
         "System redirects to product catalog with message 'Your cart is empty'",
         "Medium", "Not Run"],
        ["TC-211", "REQ-203", "Installment calculator accuracy",
         "Product price KES 50,000; interest rate 15% APR",
         "1. Select product\n2. Open installment calculator\n3. Compare 3, 6, 12, 18 month plans",
         "Monthly amounts calculated correctly with interest; total payable amount matches manual calculation",
         "Medium", "Not Run"],
        ["TC-212", "REQ-207", "Order history with filter and search",
         "Customer has 15 orders across different statuses and dates",
         "1. Open Order History\n2. Filter by 'Delivered'\n3. Search by order number\n4. Filter by date range",
         "Filters return correct results; search finds exact order; date range filter works correctly",
         "Low", "Not Run"],
    ]
    center = {1, 7, 8}
    for idx, case in enumerate(cases):
        add_data_row(ws, case, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 7: User Stories – Delivery Dashboard ───────────────────────────────

def gen_file_07():
    path = os.path.join(BASE_DIR, "03-delivery-dashboard", "User_Stories_Delivery.xlsx")
    wb, ws = create_workbook("User Stories")
    headers = [
        "ID", "Epic", "As a (Role)", "I want to (Action)",
        "So that (Benefit)", "Acceptance Criteria (Given/When/Then)",
        "Priority", "Story Points", "Status"
    ]
    col_widths = [10, 20, 22, 30, 30, 50, 12, 14, 12]
    add_header_row(ws, headers, col_widths)

    stories = [
        ["US-301", "Tracking", "Operations Manager",
         "view real-time GPS location of all delivery agents",
         "I can monitor delivery progress and optimize routes",
         "Given I am logged into the delivery dashboard\nWhen I open the live map view\nThen I see all active delivery agents with real-time GPS positions and route overlays",
         "High", 8, "Ready"],
        ["US-302", "Delivery Process", "Delivery Agent",
         "log delivery checkpoints (departed, arrived, delivered)",
         "the delivery timeline is accurately recorded",
         "Given I am assigned a delivery\nWhen I tap 'Log Checkpoint' and select status\nThen the checkpoint is recorded with GPS coordinates and timestamp",
         "High", 5, "Ready"],
        ["US-303", "Delivery Process", "Delivery Agent",
         "capture customer e-signature upon delivery",
         "I have proof of delivery",
         "Given I mark an order as delivered\nWhen the customer signs on the device screen\nThen the signature is saved to the order record with timestamp",
         "High", 5, "Ready"],
        ["US-304", "Status Updates", "Customer",
         "receive real-time status updates for my delivery",
         "I know when to expect my order",
         "Given my order is out for delivery\nWhen the agent logs each checkpoint\nThen I receive SMS with updated status and estimated delivery time",
         "Medium", 3, "Ready"],
        ["US-305", "Alerts", "Operations Manager",
         "receive alerts for delivery exceptions and delays",
         "I can take corrective action quickly",
         "Given a delivery is delayed beyond SLA threshold\nWhen the system detects the delay\nThen an alert is sent to the operations manager with agent ID and order details",
         "High", 5, "Ready"],
        ["US-306", "Analytics", "Operations Manager",
         "view delivery performance analytics",
         "I can identify bottlenecks and improve efficiency",
         "Given I open the analytics page\nWhen I select a date range\nThen I see delivery success rate, average delivery time, agent performance rankings, and SLA compliance",
         "Medium", 8, "Ready"],
        ["US-307", "Agent Management", "Operations Manager",
         "assign and reassign deliveries to agents",
         "I can balance workload across the team",
         "Given I have pending deliveries and available agents\nWhen I drag a delivery to an agent on the dashboard\nThen the assignment is updated and the agent receives notification",
         "High", 5, "Ready"],
        ["US-308", "Notifications", "Customer",
         "receive SMS notification when my order is delivered",
         "I have confirmation that delivery is complete",
         "Given the delivery agent marks order as delivered\nWhen the e-signature is captured\nThen an SMS is sent to the customer with delivery confirmation and time",
         "Medium", 3, "Ready"],
        ["US-309", "History", "Operations Manager",
         "search and filter delivery history",
         "I can look up past deliveries for dispute resolution",
         "Given I am on the delivery history page\nWhen I search by order number, date range, agent, or status\nThen matching delivery records are displayed with full timeline",
         "Low", 3, "Backlog"],
        ["US-310", "Tracking", "Operations Manager",
         "filter active deliveries by zone, agent, or status",
         "I can focus on specific areas or issues",
         "Given there are 50 active deliveries\nWhen I apply zone and status filters\nThen only matching deliveries appear on the map and list view",
         "Low", 3, "Backlog"],
    ]
    center = {1, 7, 8, 9}
    for idx, story in enumerate(stories):
        add_data_row(ws, story, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 8: RICE – Delivery Dashboard ───────────────────────────────────────

def gen_file_08():
    path = os.path.join(BASE_DIR, "03-delivery-dashboard", "RICE_Delivery.xlsx")
    wb, ws = create_workbook("RICE Prioritization")
    headers = [
        "ID", "Requirement", "Reach (users/month)", "Impact (1-3)",
        "Confidence (%)", "Effort (person-weeks)", "RICE Score (R×I×C/E)", "Priority Rank"
    ]
    col_widths = [10, 35, 20, 14, 16, 22, 24, 14]
    add_header_row(ws, headers, col_widths)

    reqs = [
        ["REQ-301", "Real-time GPS tracking map", 20, 3, 0.80, 8],
        ["REQ-302", "Delivery checkpoint logging", 50, 3, 0.90, 3],
        ["REQ-303", "E-signature capture", 50, 2, 0.85, 2],
        ["REQ-304", "Delivery performance analytics", 15, 2, 0.80, 5],
        ["REQ-305", "Exception and delay alerts", 20, 2, 0.85, 2],
        ["REQ-306", "Agent assignment management", 20, 2, 0.80, 4],
        ["REQ-307", "Customer SMS notifications", 2000, 1, 0.95, 2],
    ]
    for r in reqs:
        rice = (r[2] * r[3] * r[4]) / r[5]
        r.append(round(rice, 1))
    ranked = sorted(reqs, key=lambda x: x[6], reverse=True)
    for i, r in enumerate(ranked, 1):
        r.append(i)

    center = {1, 3, 4, 5, 6, 7, 8}
    for idx, row in enumerate(ranked):
        add_data_row(ws, row, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 9: UAT Test Plan – Delivery Dashboard ─────────────────────────────

def gen_file_09():
    path = os.path.join(BASE_DIR, "03-delivery-dashboard", "UAT_Test_Plan_Delivery.xlsx")
    wb, ws = create_workbook("UAT Test Cases")
    headers = [
        "TC-ID", "Requirement Ref", "Test Scenario", "Pre-conditions",
        "Test Steps", "Expected Result", "Priority", "Status"
    ]
    col_widths = [10, 18, 28, 28, 40, 40, 12, 12]
    add_header_row(ws, headers, col_widths)

    cases = [
        ["TC-301", "REQ-301", "Real-time GPS tracking display",
         "5 delivery agents active with GPS enabled",
         "1. Login to delivery dashboard\n2. Open live map view\n3. Verify agent positions",
         "All 5 agents visible on map with correct real-time positions; positions update every 10 seconds",
         "Critical", "Not Run"],
        ["TC-302", "REQ-302", "Log delivery checkpoint - departed",
         "Agent assigned delivery; at pickup location",
         "1. Agent opens delivery detail\n2. Taps 'Log Checkpoint'\n3. Selects 'Departed'\n4. Confirms",
         "Checkpoint recorded with GPS coordinates, timestamp; status updated; customer notified",
         "Critical", "Not Run"],
        ["TC-303", "REQ-302", "Log delivery checkpoint - delivered",
         "Agent at customer location with correct order",
         "1. Agent marks as 'Arrived'\n2. Captures e-signature\n3. Marks 'Delivered'",
         "All checkpoints recorded; e-signature saved; order status = Delivered; confirmation SMS sent",
         "Critical", "Not Run"],
        ["TC-304", "REQ-303", "E-signature capture and storage",
         "Agent at delivery point; customer present",
         "1. Agent taps 'Capture Signature'\n2. Customer signs on screen\n3. Agent confirms",
         "Signature image stored in order record; timestamp recorded; signature visible in order detail view",
         "High", "Not Run"],
        ["TC-305", "REQ-305", "Delivery delay alert trigger",
         "Delivery SLA is 2 hours; agent has not logged checkpoint for 2.5 hours",
         "1. Simulate delay scenario\n2. Wait for alert system to detect\n3. Check operations manager notification",
         "Alert sent to operations manager with agent ID, order number, and delay duration",
         "High", "Not Run"],
        ["TC-306", "REQ-304", "Performance analytics accuracy",
         "50 deliveries completed in last 7 days; mixed success/failure",
         "1. Open analytics page\n2. Select last 7 days\n3. Compare metrics with manual count",
         "Success rate, average time, and SLA compliance match manual calculations",
         "Medium", "Not Run"],
        ["TC-307", "REQ-306", "Agent assignment from dashboard",
         "3 pending deliveries; 5 agents available",
         "1. Open agent assignment view\n2. Drag delivery to agent\n3. Verify assignment",
         "Delivery reassigned; agent receives push notification; previous agent notified of removal",
         "High", "Not Run"],
        ["TC-308", "REQ-307", "Customer SMS on delivery",
         "Order assigned to agent; customer phone number valid",
         "1. Agent logs 'Out for Delivery'\n2. Agent logs 'Delivered'\n3. Check SMS logs",
         "SMS sent at each checkpoint; delivery SMS includes confirmation and time",
         "High", "Not Run"],
        ["TC-309", "REQ-301", "Filter deliveries by zone",
         "Deliveries across 4 zones active",
         "1. Open map view\n2. Apply zone filter 'Zone A'\n3. Verify displayed deliveries",
         "Only Zone A deliveries visible on map and list; other zones hidden",
         "Medium", "Not Run"],
        ["TC-310", "REQ-302", "GPS coordinates accuracy on checkpoint",
         "Agent at known location; GPS enabled",
         "1. Agent logs checkpoint\n2. System captures GPS\n3. Compare with actual location",
         "GPS coordinates within 50 meters of actual location; timestamp accurate to the second",
         "Medium", "Not Run"],
        ["TC-311", "REQ-306", "Reassign delivery to different agent",
         "Delivery assigned to Agent A; Agent B available",
         "1. Manager opens delivery detail\n2. Clicks 'Reassign'\n3. Selects Agent B\n4. Confirms",
         "Delivery moved to Agent B; Agent A notified of removal; Agent B notified of new assignment",
         "High", "Not Run"],
        ["TC-312", "REQ-304", "Date range filter on analytics",
         "90 days of delivery data available",
         "1. Open analytics\n2. Select last 30 days\n3. Then select custom range\n4. Verify data changes",
         "Metrics update correctly for each date range; no data leakage between ranges",
         "Low", "Not Run"],
    ]
    center = {1, 7, 8}
    for idx, case in enumerate(cases):
        add_data_row(ws, case, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 10: User Stories – Credit Lifecycle ────────────────────────────────

def gen_file_10():
    path = os.path.join(BASE_DIR, "04-credit-lifecycle", "User_Stories_Lifecycle.xlsx")
    wb, ws = create_workbook("User Stories")
    headers = [
        "ID", "Epic", "As a (Role)", "I want to (Action)",
        "So that (Benefit)", "Acceptance Criteria (Given/When/Then)",
        "Priority", "Story Points", "Status"
    ]
    col_widths = [10, 20, 22, 30, 30, 50, 12, 14, 12]
    add_header_row(ws, headers, col_widths)

    stories = [
        ["US-401", "Application", "Customer",
         "submit a loan application with all required documents",
         "my application can be processed without delays",
         "Given I am a registered customer\nWhen I complete the application form and upload ID, payslip, and bank statements\nThen the application is submitted and assigned an application reference number",
         "High", 8, "Ready"],
        ["US-402", "Underwriting", "Credit Analyst",
         "view all applicant documents in a unified workspace",
         "I can efficiently review applications",
         "Given an application is in 'Under Review' status\nWhen I open the underwriting workspace\nThen I see all uploaded documents, credit score, existing obligations, and application form data",
         "High", 8, "Ready"],
        ["US-403", "Underwriting", "Credit Analyst",
         "approve or reject an application with comments",
         "my decision is recorded with justification",
         "Given I have reviewed all application details\nWhen I click 'Approve' or 'Reject' and enter comments\nThen the decision is saved with my user ID, timestamp, and comments",
         "High", 5, "Ready"],
        ["US-404", "Approval", "Credit Committee",
         "review and approve applications above threshold amount",
         "high-value loans receive proper governance",
         "Given an application amount exceeds KES 500,000\nWhen the committee reviews and votes\nThen the application requires majority approval and the decision is recorded with all votes",
         "High", 8, "Ready"],
        ["US-405", "Disbursement", "Finance Officer",
         "initiate disbursement to approved customer bank account",
         "the customer receives their funds promptly",
         "Given an application is fully approved\nWhen I click 'Disburse' and confirm bank details\nThen funds are transferred within 24 hours and customer is notified",
         "High", 5, "Ready"],
        ["US-406", "Repayment", "Customer",
         "view my loan balance and make payments via mobile",
         "I can manage my loan repayments conveniently",
         "Given I have an active loan\nWhen I open 'My Loan' in the app\nThen I see outstanding balance, next due date, minimum payment, and a 'Pay Now' button",
         "High", 5, "Ready"],
        ["US-407", "Collection", "Collection Officer",
         "view list of overdue accounts with contact information",
         "I can prioritize follow-up actions",
         "Given accounts are past their due dates\nWhen I open the collection queue\nThen I see all overdue accounts sorted by days overdue with last contact date and amount due",
         "High", 5, "Ready"],
        ["US-408", "Overdue Management", "Collection Officer",
         "record collection attempt outcomes",
         "the collection history is tracked for each account",
         "Given I contact a customer about overdue payment\nWhen I log the attempt with outcome (promised to pay, unreachable, disputed)\nThen the record is saved with date, time, and notes",
         "Medium", 3, "Ready"],
        ["US-409", "Customer Portal", "Customer",
         "download my loan statements and payment receipts",
         "I have records for my personal finance management",
         "Given I have an active or completed loan\nWhen I navigate to 'Statements' and select a period\nThen I can download PDF statement with full transaction history",
         "Medium", 3, "Ready"],
        ["US-410", "Reporting", "Risk Manager",
         "view portfolio-level credit risk reports",
         "I can monitor overall portfolio health",
         "Given I access the credit risk dashboard\nWhen I view the portfolio summary\nThen I see total disbursed, outstanding portfolio, PAR >30, NPL ratio, and provisioning requirements",
         "High", 8, "Ready"],
        ["US-411", "Notifications", "Customer",
         "receive payment reminders before due date",
         "I don't miss my payment deadlines",
         "Given my payment is due in 3 days\nWhen the reminder scheduler triggers\nThen I receive an SMS with amount due, due date, and payment instructions",
         "Medium", 3, "Ready"],
        ["US-412", "Notifications", "Customer",
         "receive immediate notification when payment is posted",
         "I have confirmation that my payment was received",
         "Given I make a payment via mobile money\nWhen the payment is posted to my loan account\nThen I receive SMS confirmation with new balance and next due date",
         "Medium", 2, "Ready"],
    ]
    center = {1, 7, 8, 9}
    for idx, story in enumerate(stories):
        add_data_row(ws, story, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 11: RICE – Credit Lifecycle ────────────────────────────────────────

def gen_file_11():
    path = os.path.join(BASE_DIR, "04-credit-lifecycle", "RICE_Lifecycle.xlsx")
    wb, ws = create_workbook("RICE Prioritization")
    headers = [
        "ID", "Requirement", "Reach (users/month)", "Impact (1-3)",
        "Confidence (%)", "Effort (person-weeks)", "RICE Score (R×I×C/E)", "Priority Rank"
    ]
    col_widths = [10, 35, 20, 14, 16, 22, 24, 14]
    add_header_row(ws, headers, col_widths)

    reqs = [
        ["REQ-401", "Loan application and document management", 300, 3, 0.85, 8],
        ["REQ-402", "Automated credit scoring integration", 300, 3, 0.90, 4],
        ["REQ-403", "Approval workflow with committee voting", 100, 3, 0.80, 6],
        ["REQ-404", "Disbursement to bank account", 200, 3, 0.85, 4],
        ["REQ-405", "Mobile repayment processing", 500, 3, 0.90, 5],
        ["REQ-406", "Collection management queue", 30, 2, 0.80, 3],
        ["REQ-407", "Portfolio risk reporting dashboard", 10, 2, 0.75, 5],
        ["REQ-408", "Customer notification engine", 500, 1, 0.95, 2],
    ]
    for r in reqs:
        rice = (r[2] * r[3] * r[4]) / r[5]
        r.append(round(rice, 1))
    ranked = sorted(reqs, key=lambda x: x[6], reverse=True)
    for i, r in enumerate(ranked, 1):
        r.append(i)

    center = {1, 3, 4, 5, 6, 7, 8}
    for idx, row in enumerate(ranked):
        add_data_row(ws, row, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── File 12: UAT Test Plan – Credit Lifecycle ───────────────────────────────

def gen_file_12():
    path = os.path.join(BASE_DIR, "04-credit-lifecycle", "UAT_Test_Plan_Lifecycle.xlsx")
    wb, ws = create_workbook("UAT Test Cases")
    headers = [
        "TC-ID", "Requirement Ref", "Test Scenario", "Pre-conditions",
        "Test Steps", "Expected Result", "Priority", "Status"
    ]
    col_widths = [10, 18, 28, 28, 40, 40, 12, 12]
    add_header_row(ws, headers, col_widths)

    cases = [
        ["TC-401", "REQ-401", "Complete loan application submission",
         "Customer registered with valid KYC documents",
         "1. Login to customer portal\n2. Fill loan application form\n3. Upload ID, payslip, bank statement\n4. Submit",
         "Application submitted with reference number; all documents attached; status = 'Submitted'",
         "Critical", "Not Run"],
        ["TC-402", "REQ-401", "Application with missing documents",
         "Customer registered; missing bank statement",
         "1. Fill application form\n2. Upload only ID and payslip\n3. Attempt to submit",
         "System highlights missing required document; submission blocked until all documents uploaded",
         "High", "Not Run"],
        ["TC-403", "REQ-402", "Automated scoring for new application",
         "Application submitted with complete data; bureau API available",
         "1. Application enters scoring queue\n2. Scoring engine processes\n3. Check scoring result",
         "Score calculated within 10 seconds; decision (approved/referred/rejected) assigned; factors recorded",
         "Critical", "Not Run"],
        ["TC-404", "REQ-403", "Committee approval for high-value loan",
         "Application amount KES 800,000; 3 committee members active",
         "1. Application routed to committee\n2. Each member reviews and votes\n3. Majority approves",
         "Application approved when 2/3 members approve; decision recorded with all votes and comments",
         "Critical", "Not Run"],
        ["TC-405", "REQ-403", "Committee rejection with reason",
         "Application amount KES 600,000; 3 committee members active",
         "1. Application routed to committee\n2. 2 members reject with comments\n3. Check decision",
         "Application rejected; customer notified with reason summary; application closed",
         "Critical", "Not Run"],
        ["TC-406", "REQ-404", "Disbursement to customer bank account",
         "Application fully approved; customer bank account verified",
         "1. Finance officer initiates disbursement\n2. Confirms bank details\n3. Submits for transfer\n4. Check transfer status",
         "Transfer initiated within 24 hours; customer notified; disbursement record created in loan ledger",
         "Critical", "Not Run"],
        ["TC-407", "REQ-404", "Disbursement with incorrect bank details",
         "Application approved; customer bank account details changed recently",
         "1. Initiate disbursement\n2. System validates bank details\n3. Detects mismatch with KYC records",
         "Disbursement blocked; finance officer alerted to verify bank details; customer asked to confirm",
         "High", "Not Run"],
        ["TC-408", "REQ-405", "Customer makes repayment via mobile money",
         "Active loan with upcoming due date; customer has mobile money",
         "1. Customer opens 'My Loan'\n2. Taps 'Pay Now'\n3. Enters amount\n4. Completes mobile money payment",
         "Payment posted to loan account; balance updated; receipt generated; confirmation SMS sent",
         "Critical", "Not Run"],
        ["TC-409", "REQ-406", "Overdue account appears in collection queue",
         "Loan payment 5 days past due date",
         "1. Check collection queue after due date passes\n2. Verify account details displayed",
         "Overdue account appears in queue with days overdue, amount due, customer contact, and last payment date",
         "High", "Not Run"],
        ["TC-410", "REQ-406", "Record collection attempt outcome",
         "Overdue account in collection queue; officer logged in",
         "1. Open overdue account\n2. Click 'Log Contact'\n3. Select outcome 'Promised to Pay'\n4. Enter callback date\n5. Save",
         "Contact record saved with outcome, notes, and callback date; account flagged for follow-up",
         "High", "Not Run"],
        ["TC-411", "REQ-407", "Portfolio risk report accuracy",
         "Portfolio with 100 active loans; known NPL ratio",
         "1. Open portfolio dashboard\n2. Compare displayed NPL ratio with manual calculation\n3. Check PAR metrics",
         "All portfolio metrics (PAR30, PAR60, PAR90, NPL ratio) match manual calculations within 0.1% tolerance",
         "Medium", "Not Run"],
        ["TC-412", "REQ-408", "Payment reminder sent 3 days before due",
         "Active loan with due date in 3 days; customer phone valid",
         "1. Scheduler triggers 3-day reminder\n2. Check SMS delivery logs",
         "SMS delivered with correct amount, due date, and payment instructions; no duplicate messages",
         "High", "Not Run"],
        ["TC-413", "REQ-405", "Partial payment handling",
         "Active loan; minimum payment KES 5,000; customer pays KES 3,000",
         "1. Customer initiates payment for KES 3,000\n2. Payment processed\n3. Check account status",
         "Partial payment posted; remaining balance still due; next reminder scheduled; no penalty waived automatically",
         "Medium", "Not Run"],
        ["TC-414", "REQ-409", "Customer downloads loan statement",
         "Active loan with 6 months payment history",
         "1. Customer navigates to 'Statements'\n2. Selects last 6 months\n3. Downloads PDF",
         "PDF statement downloaded with all transactions, running balance, and interest breakdown for the period",
         "Medium", "Not Run"],
        ["TC-415", "REQ-404", "Disbursement retry on bank transfer failure",
         "Disbursement initiated; bank returns failure response",
         "1. Disbursement submitted\n2. Bank API returns failure\n3. System queues for retry\n4. Retry succeeds on 2nd attempt",
         "First failure logged; automatic retry after 1 hour; successful disbursement on retry; all events recorded",
         "High", "Not Run"],
    ]
    center = {1, 7, 8}
    for idx, case in enumerate(cases):
        add_data_row(ws, case, idx + 2, center_cols=center)
    save_workbook(wb, path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  Embafinans BA Practice Artifacts - XLSX Generator")
    print("=" * 70)
    print()

    generators = [
        ("File  1: User Stories - Credit Scoring",       gen_file_01),
        ("File  2: RICE - Credit Scoring",               gen_file_02),
        ("File  3: UAT Test Plan - Credit Scoring",      gen_file_03),
        ("File  4: User Stories - B2C Sales Channel",    gen_file_04),
        ("File  5: RICE - B2C Sales Channel",            gen_file_05),
        ("File  6: UAT Test Plan - B2C Sales Channel",   gen_file_06),
        ("File  7: User Stories - Delivery Dashboard",   gen_file_07),
        ("File  8: RICE - Delivery Dashboard",           gen_file_08),
        ("File  9: UAT Test Plan - Delivery Dashboard",  gen_file_09),
        ("File 10: User Stories - Credit Lifecycle",     gen_file_10),
        ("File 11: RICE - Credit Lifecycle",             gen_file_11),
        ("File 12: UAT Test Plan - Credit Lifecycle",    gen_file_12),
    ]

    success_count = 0
    fail_count = 0

    for label, gen_func in generators:
        try:
            gen_func()
            success_count += 1
        except Exception as e:
            print(f"  ✗ FAILED: {label} — {e}", file=sys.stderr)
            fail_count += 1

    print()
    print("-" * 70)
    print(f"  Results: {success_count} succeeded, {fail_count} failed out of {len(generators)}")
    print("-" * 70)
    return fail_count == 0


if __name__ == "__main__":
    ok = main()
    sys.exit(0 if ok else 1)
