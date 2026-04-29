import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'skills', 'xlsx', 'templates'))
from base import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()
wb.properties.creator = "Z.ai"

def make_sheet(ws, title, headers, sample_data, extra_rows, col_widths=None):
    """Helper to create a formatted sheet."""
    ws.sheet_view.showGridLines = False
    last_col = len(headers) + 1
    setup_sheet(ws, title=title, last_col=last_col)

    for col_idx, h in enumerate(headers, start=2):
        ws.cell(row=4, column=col_idx, value=h)
    style_header_row(ws, row_num=4, col_start=2, col_end=last_col)

    for i, row_data in enumerate(sample_data):
        row_num = 5 + i
        for col_idx, val in enumerate(row_data, start=2):
            ws.cell(row=row_num, column=col_idx, value=val)
        style_data_row(ws, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

    for i in range(len(sample_data), len(sample_data) + extra_rows):
        row_num = 5 + i
        for col_idx in range(2, last_col + 1):
            ws.cell(row=row_num, column=col_idx, value="")
        style_data_row(ws, row_num=row_num, col_start=2, col_end=last_col, row_index=i)

    ws.freeze_panes = 'B5'
    if col_widths:
        for idx, w in enumerate(col_widths, start=2):
            ws.column_dimensions[get_column_letter(idx)].width = w
    else:
        auto_fit_columns(ws, min_width=10, max_width=35, header_row=4, data_start_row=5)

    return last_col

def add_status_cf(ws, col_letter, start_row, end_row):
    """Add conditional formatting for status columns."""
    rng = f'{col_letter}{start_row}:{col_letter}{end_row}'
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"Done"'], fill=PatternFill(bgColor='E8F5E9'), font=Font(color=ACCENT_POSITIVE)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"In Progress"'], fill=PatternFill(bgColor='FEF9E7'), font=Font(color=ACCENT_WARNING)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"Blocked"'], fill=PatternFill(bgColor='FDEDEC'), font=Font(color=ACCENT_NEGATIVE)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"Overdue"'], fill=PatternFill(bgColor='FDEDEC'), font=Font(color=ACCENT_NEGATIVE)))
    ws.conditional_formatting.add(rng,
        CellIsRule(operator='equal', formula=['"Delayed"'], fill=PatternFill(bgColor='FDEDEC'), font=Font(color=ACCENT_NEGATIVE)))

# ============================================================
# TAB 1: REQUIREMENTS SUMMARY
# ============================================================
ws1 = wb.active
ws1.title = "Requirements"

headers1 = ["BR No.", "Requirement Title", "Priority", "Status", "Sprint", "Assigned Developer", "Linked US", "Linked AC", "Notes"]
sample_reqs = [
    ["BR-001", "Online Return Request Submission", "High", "Done", "Sprint 1", "Tural", "US-001 to US-003", "AC-001 to AC-005", ""],
    ["BR-002", "Automated Eligibility Check", "High", "Done", "Sprint 1", "Tural", "US-004", "AC-006 to AC-009", "14-day policy + product condition"],
    ["BR-003", "Digital Approval Workflow", "High", "In Progress", "Sprint 3", "Tural", "US-005 to US-007", "AC-010 to AC-015", "Auto-approve under 100 AZN"],
    ["BR-004", "Customer Return Status Tracking", "Medium", "In Progress", "Sprint 3", "Rashad", "US-008 to US-009", "AC-016 to AC-019", "Email + SMS notifications"],
    ["BR-005", "Warehouse Pickup Notification", "Medium", "Not Started", "Sprint 4", "Tural", "US-010", "AC-020 to AC-022", ""],
    ["BR-006", "Refund Processing Integration", "High", "Not Started", "Sprint 4", "Rashad", "US-011 to US-013", "AC-023 to AC-028", "Finance team dependency"],
]
lc1 = make_sheet(ws1, "Requirements Summary", headers1, sample_reqs, 20)

dv1 = DataValidation(type="list", formula1='"Not Started,In Progress,Testing,Done,Blocked"', allow_blank=True)
ws1.add_data_validation(dv1)
dv1.add('E5:E50')

dv1b = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws1.add_data_validation(dv1b)
dv1b.add('D5:D50')

add_status_cf(ws1, 'E', 5, 50)

# ============================================================
# TAB 2: SPRINT BOARD
# ============================================================
ws2 = wb.create_sheet("Sprint Board")
headers2 = ["Sprint", "Sprint Goal", "Stories", "Status", "Start Date", "End Date", "Blockers", "Expected Delivery", "Actual Delivery", "Notes"]
sample_sprints = [
    ["Sprint 1", "Online return request + eligibility check", "US-001 to US-004", "Done", "2026-04-06", "2026-04-17", "None", "2026-04-17", "2026-04-17", "On time"],
    ["Sprint 2", "Database schema + API setup", "US-005 (partial)", "Done", "2026-04-20", "2026-05-01", "None", "2026-05-01", "2026-05-01", "DB design took extra day"],
    ["Sprint 3", "Approval workflow + status tracking", "US-005 to US-009", "In Progress", "2026-05-04", "2026-05-15", "Notification service delay", "2026-05-15", "", "Waiting for SMS provider"],
    ["Sprint 4", "Warehouse notification + refund integration", "US-010 to US-013", "Not Started", "2026-05-18", "2026-05-29", "", "2026-05-29", "", ""],
    ["Sprint 5", "UAT + bug fixes + production release", "US-014+", "Not Started", "2026-06-01", "2026-06-12", "", "2026-06-12", "", ""],
]
lc2 = make_sheet(ws2, "Sprint Board", headers2, sample_sprints, 10)

dv2 = DataValidation(type="list", formula1='"Not Started,In Progress,Done,Delayed"', allow_blank=True)
ws2.add_data_validation(dv2)
dv2.add('E5:E50')
add_status_cf(ws2, 'E', 5, 50)

# ============================================================
# TAB 3: ACTION ITEMS
# ============================================================
ws3 = wb.create_sheet("Action Items")
headers3 = ["Date", "Description", "Owner", "Deadline", "Status", "Priority", "Related BR/US", "Follow-up Date", "Notes"]
sample_actions = [
    ["2026-04-28", "Update BR-003 approval threshold from 200 to 100 AZN", "Zamir", "2026-04-29", "Done", "High", "BR-003, US-005", "2026-04-29", "Elnar approved"],
    ["2026-04-29", "Clarify AC-012 photo size limit with Aysel", "Zamir", "2026-04-30", "Done", "High", "AC-012", "2026-04-30", "Max 5 MB confirmed"],
    ["2026-05-01", "Review API documentation from Tural", "Zamir", "2026-05-04", "In Progress", "High", "BR-001 to BR-003", "2026-05-04", ""],
    ["2026-05-02", "Schedule UAT kickoff meeting with stakeholders", "Zamir", "2026-05-15", "Not Started", "Medium", "All", "", "Depends on Sprint 3"],
    ["2026-05-02", "Get SMS provider options from IT department", "Ramin", "2026-05-06", "In Progress", "High", "BR-004", "2026-05-06", "Sprint 3 blocker"],
]
lc3 = make_sheet(ws3, "Action Items", headers3, sample_actions, 45)

dv3a = DataValidation(type="list", formula1='"Not Started,In Progress,Done,Overdue"', allow_blank=True)
ws3.add_data_validation(dv3a)
dv3a.add('F5:F200')
dv3b = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws3.add_data_validation(dv3b)
dv3b.add('G5:G200')
add_status_cf(ws3, 'F', 5, 200)

# ============================================================
# TAB 4: DECISIONS LOG
# ============================================================
ws4 = wb.create_sheet("Decisions Log")
headers4 = ["Date", "Decision", "Requested By", "Discussed With", "Rationale", "Impact", "Related BR/US", "Status", "Notes"]
sample_decisions = [
    ["2026-04-10", "Mobile app will NOT be built in V1", "Elnar (Sponsor)", "Ramin, Tural, Zamir", "Budget constraint (80K AZN) and 3-month timeline", "Reduced scope, web-only", "Scope", "Approved", "Documented in BRD section 4"],
    ["2026-04-12", "Auto-approve returns under 100 AZN", "Tural (Lead Dev)", "Ramin, Zamir", "Reduces approval bottleneck, 70% under 100 AZN", "Simplified workflow", "BR-003", "Approved", ""],
    ["2026-04-15", "Photo upload optional for returns", "Narmin (Support)", "Zamir, Aysel", "Customers may not have photos", "Max 5 MB, JPG/PNG only", "US-001", "Approved", "Aysel updated AC-003"],
    ["2026-04-22", "Notifications via Email + SMS", "Zamir", "Tural, Ramin", "Customers need real-time updates", "Added SMS provider dependency", "BR-004", "Approved", "Sprint 3 blocker"],
    ["2026-04-28", "Return policy 14 to 30 days (change request)", "Elnar (Sponsor)", "Ramin, Tural, Zamir", "Marketing team feedback", "Affects 4 US, 8 AC, +2 dev days", "BR-001, BR-002", "Pending", "Needs sprint plan update"],
]
lc4 = make_sheet(ws4, "Decisions Log", headers4, sample_decisions, 25)

dv4 = DataValidation(type="list", formula1='"Approved,Pending,Rejected,Deferred"', allow_blank=True)
ws4.add_data_validation(dv4)
dv4.add('I5:I200')

ws4.conditional_formatting.add('I5:I200',
    CellIsRule(operator='equal', formula=['"Approved"'], fill=PatternFill(bgColor='E8F5E9'), font=Font(color=ACCENT_POSITIVE)))
ws4.conditional_formatting.add('I5:I200',
    CellIsRule(operator='equal', formula=['"Pending"'], fill=PatternFill(bgColor='FEF9E7'), font=Font(color=ACCENT_WARNING)))
ws4.conditional_formatting.add('I5:I200',
    CellIsRule(operator='equal', formula=['"Rejected"'], fill=PatternFill(bgColor='FDEDEC'), font=Font(color=ACCENT_NEGATIVE)))

# ============================================================
# TAB 5: RISK REGISTER
# ============================================================
ws5 = wb.create_sheet("Risk Register")
headers5 = ["Risk ID", "Risk Description", "Impact", "Probability", "Risk Score", "Mitigation Plan", "Owner", "Status", "Notes"]
sample_risks = [
    ["RSK-001", "SMS provider not selected by Sprint 3 start", "High", "Medium", 6, "Ask Ramin to escalate to IT by April 30", "Ramin", "Open", "Sprint 3 blocker"],
    ["RSK-002", "Area manager unavailable for approval (2-3 hr wait)", "Medium", "High", 6, "Auto-approve under 100 AZN (reduces 70%)", "Tural", "Mitigated", "Decision D-002"],
    ["RSK-003", "Finance team cannot complete UAT on schedule", "High", "Low", 3, "Schedule UAT 2 weeks in advance", "Zamir", "Open", ""],
    ["RSK-004", "Scope creep from stakeholder change requests", "Medium", "High", 6, "Formal change request process", "Zamir", "Open", "Already happened once"],
    ["RSK-005", "Database performance with large return volumes", "Medium", "Low", 3, "Tural to run load tests in Sprint 4", "Tural", "Open", ""],
]
lc5 = make_sheet(ws5, "Risk Register", headers5, sample_risks, 15)

dv5a = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
dv5b = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
dv5c = DataValidation(type="list", formula1='"Open,Mitigated,Closed,Accepted"', allow_blank=True)
ws5.add_data_validation(dv5a)
ws5.add_data_validation(dv5b)
ws5.add_data_validation(dv5c)
dv5a.add('D5:D100')
dv5b.add('E5:E100')
dv5c.add('I5:I100')

ws5.conditional_formatting.add('F5:F100',
    CellIsRule(operator='greaterThanOrEqual', formula=['6'], fill=PatternFill(bgColor='FDEDEC'), font=Font(color=ACCENT_NEGATIVE, bold=True)))
ws5.conditional_formatting.add('F5:F100',
    CellIsRule(operator='between', formula=['3', '5'], fill=PatternFill(bgColor='FEF9E7'), font=Font(color=ACCENT_WARNING)))
ws5.conditional_formatting.add('F5:F100',
    CellIsRule(operator='lessThanOrEqual', formula=['2'], fill=PatternFill(bgColor='E8F5E9'), font=Font(color=ACCENT_POSITIVE)))

# ============================================================
# TAB 6: STAKEHOLDER MAP
# ============================================================
ws6 = wb.create_sheet("Stakeholder Map")
headers6 = ["Name", "Role", "Department", "Email / Phone", "Interest", "Influence", "Key Concerns", "Comm. Frequency", "Notes"]
sample_sh = [
    ["Elnar Huseynov", "Head of Retail / Sponsor", "Retail", "elnar@kontakt.az", "High", "High", "Budget, timeline, customer satisfaction", "Weekly", "Approves scope and budget"],
    ["Ramin Aliyev", "Project Manager", "PMO", "ramin@kontakt.az", "High", "High", "Sprint delivery, resources, risks", "Daily", "Main contact for schedule"],
    ["Tural Mammadov", "Lead Developer", "IT", "tural@kontakt.az", "Medium", "High", "Tech feasibility, architecture", "Daily", "Primary tech decision maker"],
    ["Rashad Karimov", "Backend Developer", "IT", "rashad@kontakt.az", "Medium", "Medium", "API design, database", "Daily (standup)", "Working on BR-004"],
    ["Aysel Ahmadova", "QA Tester", "IT", "aysel@kontakt.az", "Medium", "Medium", "Test coverage, clear AC", "Daily (standup)", "Reviews all AC"],
    ["Leyla Hasanova", "Store Manager", "Retail", "leyla@kontakt.az", "High", "Low", "Customer complaints, wait times", "Bi-weekly", "UAT business user"],
    ["Narmin Guliyeva", "Customer Support Lead", "Support", "narmin@kontakt.az", "High", "Medium", "Call volume, self-service", "Bi-weekly", "40% calls are returns"],
    ["Samir Abdullayev", "Warehouse Supervisor", "Logistics", "samir@kontakt.az", "Medium", "Low", "Pickup scheduling, storage", "Monthly", "Needs advance notice"],
    ["Rasul Novruzov", "Finance Officer", "Finance", "rasul@kontakt.az", "Medium", "Medium", "Refund accuracy", "Bi-weekly", "Single point of failure"],
]
lc6 = make_sheet(ws6, "Stakeholder Map", headers6, sample_sh, 10)

dv6a = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
dv6b = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
dv6c = DataValidation(type="list", formula1='"Daily,Weekly,Bi-weekly,Monthly,As needed"', allow_blank=True)
ws6.add_data_validation(dv6a)
ws6.add_data_validation(dv6b)
ws6.add_data_validation(dv6c)
dv6a.add('F5:F50')
dv6b.add('G5:G50')
dv6c.add('I5:I50')

# ============================================================
# TAB 7: MEETING NOTES
# ============================================================
ws7 = wb.create_sheet("Meeting Notes")
headers7 = ["Date", "Meeting Type", "Attendees", "Agenda", "Discussion", "Decisions", "Action Items"]
sample_mn = [
    ["2026-04-28", "Sprint Planning", "Ramin, Tural, Rashad, Aysel, Zamir", "Select stories for Sprint 3", "US-005 to US-009 selected. SMS provider concern raised.", "Auto-approve under 100 AZN confirmed", "Zamir: Review API docs by May 4. Ramin: Get SMS options by May 6."],
    ["2026-04-25", "Backlog Refinement", "Ramin, Tural, Zamir", "Review Sprint 4 stories", "US-010 needs more detail. Integration with logistics system.", "US-010 needs rewrite", "Zamir: Rewrite US-010 by May 1."],
    ["2026-04-22", "Change Request Review", "Elnar, Ramin, Zamir", "14 to 30 day policy change", "Marketing wants 30 days. Impact: 4 US, 8 AC, +2 dev days.", "Pending Ramin update", "Zamir: Update US/AC. Ramin: Update sprint plan."],
]
lc7 = make_sheet(ws7, "Meeting Notes", headers7, sample_mn, 25)

dv7 = DataValidation(type="list", formula1='"Sprint Planning,Backlog Refinement,Daily Standup,Stakeholder Interview,Change Request Review,UAT Kickoff,Retrospective,Technical Discussion,Other"', allow_blank=True)
ws7.add_data_validation(dv7)
dv7.add('C5:C100')

# ============================================================
# TAB 8: TRACEABILITY MATRIX
# ============================================================
ws8 = wb.create_sheet("Traceability")
headers8 = ["BR No.", "BR Title", "User Stories", "Acceptance Criteria", "API Endpoints", "Test Cases", "Status"]
sample_tr = [
    ["BR-001", "Online Return Request", "US-001, US-002, US-003", "AC-001 to AC-005", "POST /api/returns\nGET /api/returns/{id}\nPOST /api/returns/{id}/cancel", "TC-001 to TC-008", "Done"],
    ["BR-002", "Eligibility Check", "US-004", "AC-006 to AC-009", "GET /api/returns/eligibility/{productId}", "TC-009 to TC-012", "Done"],
    ["BR-003", "Approval Workflow", "US-005, US-006, US-007", "AC-010 to AC-015", "POST /api/approvals\nPUT /api/approvals/{id}\nGET /api/approvals/pending", "TC-013 to TC-020", "In Progress"],
    ["BR-004", "Status Tracking", "US-008, US-009", "AC-016 to AC-019", "GET /api/returns/{id}/status\nPOST /api/notifications", "TC-021 to TC-025", "In Progress"],
    ["BR-005", "Warehouse Notification", "US-010", "AC-020 to AC-022", "POST /api/warehouse/pickup\nGET /api/warehouse/schedule", "TC-026 to TC-028", "Not Started"],
    ["BR-006", "Refund Processing", "US-011, US-012, US-013", "AC-023 to AC-028", "POST /api/refunds\nGET /api/refunds/{id}\nPUT /api/refunds/{id}", "TC-029 to TC-035", "Not Started"],
]
lc8 = make_sheet(ws8, "Requirements Traceability Matrix", headers8, sample_tr, 14)

dv8 = DataValidation(type="list", formula1='"Not Started,In Progress,Testing,Done"', allow_blank=True)
ws8.add_data_validation(dv8)
dv8.add('H5:H100')
add_status_cf(ws8, 'H', 5, 100)

# ============================================================
# SAVE
# ============================================================
output_path = "/home/z/my-project/ba-practice/BA_Project_Tracker_Kontakt_Home.xlsx"
wb.save(output_path)
print(f"Workbook saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
